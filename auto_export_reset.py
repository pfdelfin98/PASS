import pymysql
import openpyxl
import os
from datetime import datetime, timedelta


DAYS_IN_WEEK = {
    0: "Monday",
    1: "Tuesday",
    2: "Wednesday",
    3: "Thursday",
    4: "Friday",
    5: "Saturday",
    6: "Sunday",
}


class AutomaticExportResetLogs(object):
    def __init__(
        self,
    ):
        self.day_of_export = "Sunday"
        self.folder_name = "Auto_Export"
        self.analyticts_folder_path = (
            rf"C:\Users\Cj\Desktop\Analytics\{self.folder_name}"  # Change path
        )
        self.logs_folder_path = (
            rf"C:\Users\Cj\Desktop\Logs\{self.folder_name}"  # Change Path
        )

    def check_if_logs_found(self, prev_monday, prev_sunday):
        connection = pymysql.connect(
            host="localhost", user="root", password="", database="pass_db"
        )
        try:
            cursor = connection.cursor()
            query = f"SELECT * FROM tbl_logs WHERE date_log BETWEEN '{prev_monday}' AND '{prev_sunday}'"
            cursor.execute(query)
            logs = cursor.fetchall()
            if len(logs) > 0:
                return True
            else:
                return False
        except Exception as e:
            print(e)
        finally:
            cursor.close
            connection.close()

    def export_logs(self, prev_monday, prev_sunday):
        # Connect to the MySQL database
        connection = pymysql.connect(
            host="localhost", user="root", password="", database="pass_db"
        )

        try:
            # Create a cursor object to execute SQL queries
            cursor = connection.cursor()

            # Retrieve data from the tbl_student table
            select_query = f"SELECT CONCAT(tbl_student.first_name, ' ', tbl_student.last_name) AS student_name, tbl_student.course, tbl_student.sr_code, tbl_logs.date_log, tbl_logs.time_log, tbl_logs.log_type FROM tbl_logs LEFT JOIN tbl_student ON tbl_logs.student_id = tbl_student.id WHERE tbl_logs.date_log BETWEEN '{prev_monday}' AND '{prev_sunday}' ORDER BY tbl_logs.date_log and tbl_logs.time_log ASC"
            cursor.execute(
                select_query,
            )
            student_data = cursor.fetchall()

            # Create a new Excel workbook and select the active sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Write the column headers
            sheet["A1"] = "Student Name"
            sheet["B1"] = "Course"
            sheet["C1"] = "SR Code"
            sheet["D1"] = "Date Log"
            sheet["E1"] = "Time Log"
            sheet["F1"] = "Log Type"

            # Set column width for date columns
            date_columns = ["D", "E"]  # Columns D and E represent the date columns
            for column in date_columns:
                sheet.column_dimensions[
                    column
                ].width = 15  # Adjust the width as per your preference

            for row_index, student in enumerate(student_data, start=2):
                sheet.cell(row=row_index, column=1).value = student[0]
                sheet.cell(row=row_index, column=2).value = student[1]
                sheet.cell(row=row_index, column=3).value = student[2]
                sheet.cell(row=row_index, column=4).value = student[3]
                sheet.cell(row=row_index, column=5).value = student[4]
                sheet.cell(row=row_index, column=6).value = student[5]

            # Save the Excel file
            self.file_name = (
                f"auto_export_logs_data_{prev_monday}_to_{prev_sunday}.xlsx"
            )

            # Save the Excel file inside the "folder_path" folder
            self.file_path = rf"{self.logs_folder_path}\{self.file_name}"

            # Create the "logs" folder if it doesn't exist
            if not os.path.exists(self.logs_folder_path):
                os.makedirs(self.logs_folder_path)
            # Save the Excel file
            workbook.save(self.file_path)
            print("Logs Data exported to Excel successfully!")

        except Exception as e:
            print("Error exporting data to Excel:", str(e))

        finally:
            # Close the cursor and connection
            cursor.close()
            connection.close()

    def export_delete_analytics(self, prev_monday, prev_sunday):
        from openpyxl.chart import BarChart, Reference

        departments = ["CABEIHM", "CAS", "CICS", "CET", "CONAHS", "CTE"]

        # Connect to the MySQL database

        connection = pymysql.connect(
            host="localhost", user="root", password="", database="pass_db"
        )
        departments_done_exporting = []
        for department in departments:
            connection.ping()
            try:
                # Execute the query and fetch the gender and gender data
                cursor = connection.cursor()
                query = f"SELECT \
                            s.course, \
                            SUM(CASE WHEN s.gender = 'male' THEN 1 ELSE 0 END) AS male_log_count, \
                            SUM(CASE WHEN s.gender = 'female' THEN 1 ELSE 0 END) AS female_log_count, \
                            SUM(CASE WHEN s.gender = 'male' OR s.gender = 'female' THEN 1 ELSE 0 END) AS total_log_count \
                        FROM \
                            tbl_student s \
                        JOIN \
                            tbl_logs l ON s.id = l.student_id \
                        WHERE s.department = '{department}' AND l.date_log BETWEEN '{prev_monday}' AND '{prev_sunday}' AND l.log_type = 'LOGGED_IN' \
                        GROUP BY s.course"
                cursor.execute(query)
                chart_data = cursor.fetchall()

                headers = ("Course", "Male", "Female", "Total Logs")
                rows = [headers]  # Column headers for excel file
                for row in chart_data:
                    rows.append(row)

                # Create a new Excel workbook and select the active sheet
                workbook = openpyxl.Workbook(write_only=True)
                sheet = workbook.create_sheet()

                for _ in rows:
                    sheet.append(_)

                # Bar Chart Config
                chart1 = BarChart()
                chart1.type = "col"
                chart1.style = 10
                chart1.title = "Dashboard"
                chart1.y_axis.title = "Count"
                chart1.x_axis.title = "Courses"
                chart1.height = 10
                chart1.width = 20

                chart_data = Reference(
                    sheet, min_col=2, min_row=1, max_row=len(rows), max_col=3
                )
                categories = Reference(sheet, min_col=1, min_row=2, max_row=len(rows))
                chart1.add_data(chart_data, titles_from_data=True)
                chart1.shape = 5
                sheet.add_chart(chart1, "F2")

                # Save the Excel file
                current_datetime = datetime.now()
                formatted_datetime = current_datetime.strftime("%Y-%m-%d_%H-%M-%S")
                self.file_name = (
                    f"Auto_Export_Analytics_{department}_{formatted_datetime}.xlsx"
                )

                # Save the Excel file inside the "folder_path" folder
                self.file_path = rf"{self.analyticts_folder_path}\Week_{prev_monday}_to_{prev_sunday}\{self.file_name}"
                file_folder_path = rf"{self.analyticts_folder_path}\Week_{prev_monday}_to_{prev_sunday}"

                # Create the  folder if it doesn't exist
                if not os.path.exists(file_folder_path):
                    os.makedirs(file_folder_path)
                # Save the Excel file
                workbook.save(self.file_path)
                print(f"Logs Data exported to Excel for {department} successfully!")

            except Exception as e:
                print("Error exporting data to Excel:", str(e))
            else:
                departments_done_exporting.append(department)

        print("Exporting data Completed!")
        cursor.close()
        connection.close()
        if len(departments_done_exporting) == len(departments):
            self.delete_analytics(prev_monday=prev_monday, prev_sunday=prev_sunday)
        else:
            print(
                "Will not delete logs! due to error in exporting departments data to excel"
            )

    def delete_analytics(self, prev_monday, prev_sunday):
        connection = pymysql.connect(
            host="localhost", user="root", password="", database="pass_db"
        )
        try:
            cursor = connection.cursor()
            query = f"DELETE FROM tbl_logs WHERE date_log BETWEEN '{prev_monday}' AND '{prev_sunday}'"
            cursor.execute(query)
            connection.commit()
            print("Logs deleted successfully!")
        except Exception as e:
            print("Error deleting logs:", str(e))
