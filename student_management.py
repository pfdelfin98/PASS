import sys
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtWidgets import (
    QTableWidget,
    QTableWidgetItem,
    QPushButton,
)
import pymysql
import os
import logs
import openpyxl
import register_student
import student_management
import admin_login
import dashboard
import archive_management
from edit_student import EditStudentDialog
from delete_student import DeleteStudentDialog
from view_logs import ViewLogsDialog
from PyQt5.QtCore import QTimer
from datetime import datetime
from archive_student import DeleteStudentDialog
from archive_management import ArchiveManagementWindow

class StudentManagementWindow(object):
    def __init__(self) -> None:
        # For Excel File
        self.file_name = ""
        self.file_path = ""
        self.folder_name = "student_logs"
        self.folder_path = rf"C:\Users\SampleUser\Desktop\{self.folder_name}"  # Change this to your own file path

    def setupUi(self, MainWindow):
        self.MainWindow = MainWindow
        MainWindow.setWindowFlags(
            MainWindow.windowFlags()
            & ~QtCore.Qt.WindowMinimizeButtonHint
            & ~QtCore.Qt.WindowMaximizeButtonHint
        )
        MainWindow.setObjectName("MainWindow")
                # MainWindow.resize(1200, 700)
        self.MainWindow.showMaximized()
        MainWindow.setWindowFlags(
            MainWindow.windowFlags()
            & ~QtCore.Qt.WindowCloseButtonHint  # Remove the close button
        )
        self.imageFilePath = None
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        # Add table widget
        self.tableWidget = QTableWidget(self.centralwidget)
        self.tableWidget.setGeometry(QtCore.QRect(310, 150, 1240, 600))
        self.tableWidget.setObjectName("tableWidget")

        self.searchLineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.searchLineEdit.setGeometry(QtCore.QRect(1050, 95, 200, 30))
        self.searchLineEdit.setObjectName("searchLineEdit")
        self.searchLineEdit.textChanged.connect(self.search_logs)
        self.search_has_input = False

        self.searchLabel = QtWidgets.QLabel(self.centralwidget)
        self.searchLabel.setGeometry(QtCore.QRect(1000, 95, 70, 30))
        self.searchLabel.setObjectName("searchLabel")
        self.searchLabel.setText("Search:")

        self.frame = QtWidgets.QFrame(self.centralwidget)
        self.frame.setGeometry(QtCore.QRect(-1, 1, 261, 2000))
        self.frame.setStyleSheet("background-color: rgb(227, 30, 36);")
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.label_4 = QtWidgets.QLabel(self.frame)
        self.label_4.setGeometry(QtCore.QRect(80, 20, 101, 91))
        self.label_4.setText("")
        self.label_4.setPixmap(QtGui.QPixmap("img/logo2.png"))
        self.label_4.setScaledContents(True)
        self.label_4.setObjectName("label_4")
        self.label_9 = QtWidgets.QLabel(self.frame)
        self.label_9.setGeometry(QtCore.QRect(44, 110, 201, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label_9.setFont(font)
        self.label_9.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_9.setObjectName("label_9")
        self.faceRecognitionBtn = QtWidgets.QPushButton(self.frame)
        self.faceRecognitionBtn.setGeometry(QtCore.QRect(40, 300, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.faceRecognitionBtn.setFont(font)
        self.faceRecognitionBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.faceRecognitionBtn.setObjectName("faceRecognitionBtn")
        self.registerStudentBtn = QtWidgets.QPushButton(self.frame)
        self.registerStudentBtn.setGeometry(QtCore.QRect(40, 350, 211, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.registerStudentBtn.setFont(font)
        self.registerStudentBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.registerStudentBtn.setObjectName("registerStudentBtn")
        self.studentMgmtBtn = QtWidgets.QPushButton(self.frame)
        self.studentMgmtBtn.setGeometry(QtCore.QRect(40, 400, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.studentMgmtBtn.setFont(font)
        self.studentMgmtBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.studentMgmtBtn.setObjectName("studentMgmtBtn")
        self.exitBtn = QtWidgets.QPushButton(self.frame)
        self.exitBtn.setGeometry(QtCore.QRect(40, 450, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.exitBtn.setFont(font)
        self.exitBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.exitBtn.setObjectName("exitBtn")
        self.exitBtn_2 = QtWidgets.QPushButton(self.frame)
        self.exitBtn_2.setGeometry(QtCore.QRect(40, 500, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.exitBtn_2.setFont(font)
        self.exitBtn_2.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.exitBtn_2.setObjectName("exitBtn_2")

        self.pushButton = QtWidgets.QPushButton(self.frame)
        self.pushButton.setGeometry(QtCore.QRect(40, 200, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.pushButton.setFont(font)
        self.pushButton.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )

        self.pushButton.setObjectName("pushButton")


        self.pushButton2 = QtWidgets.QPushButton(self.frame)
        self.pushButton2.setGeometry(QtCore.QRect(40, 250, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.pushButton2.setFont(font)
        self.pushButton2.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )

        self.pushButton2.setObjectName("pushButton2")

        # self.frame_3 = QtWidgets.QFrame(self.centralwidget)
        # self.frame_3.setGeometry(QtCore.QRect(261, -1, 2000, 61))
        # self.frame_3.setStyleSheet("background-color: rgb(255, 255, 255);")
        # self.frame_3.setFrameShape(QtWidgets.QFrame.StyledPanel)
        # self.frame_3.setFrameShadow(QtWidgets.QFrame.Raised)
        # self.frame_3.setObjectName("frame_3")
        # self.label = QtWidgets.QLabel(self.frame_3)
        # self.label.setGeometry(QtCore.QRect(20, 10, 671, 41))
        # font = QtGui.QFont()
        # font.setFamily("Arial")
        # font.setPointSize(12)
        # self.label.setFont(font)
        # self.label.setObjectName("label")
        self.label_10 = QtWidgets.QLabel(self.centralwidget)
        self.label_10.setGeometry(QtCore.QRect(310, 90, 301, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(15)
        font.setBold(True)
        font.setWeight(75)
        self.label_10.setFont(font)
        self.label_10.setStyleSheet("color:black;")
        self.label_10.setObjectName("label_10")


        archivesButton = QtWidgets.QPushButton(self.centralwidget)
        archivesButton.setGeometry(QtCore.QRect(1280, 90, 121, 40))
        archivesButton.setObjectName("archivesButton")
        archivesButton.setText("Archives")
        archivesButton.setStyleSheet(
            '''
            QPushButton {
                background-color: #dc3545;  
                border: none;
                color: white;
                padding: 8px 16px;
                border-radius: 4px;
                font-family: Arial;
                font-size: 8pt;
            }

            QPushButton:hover {
                background-color: #c82333;  
            }
            '''
        )
        archivesButton.clicked.connect(self.open_archives)

        self.exportDataBtn = QtWidgets.QPushButton(self.centralwidget)
        self.exportDataBtn.setGeometry(QtCore.QRect(1420, 90, 121, 40))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(8)
        self.exportDataBtn.setFont(font)
        self.exportDataBtn.setObjectName("exportDataBtn")
        self.exportDataBtn.setText("Export Data")
        self.exportDataBtn.setStyleSheet(
            '''
            QPushButton {
                background-color: #dc3545;  
                border: none;
                color: white;
                padding: 8px 16px;
                border-radius: 4px;
                font-family: Arial;
                font-size: 8pt;
            }

            QPushButton:hover {
                background-color: #c82333;  
            }
            '''
        )
        self.exportDataBtn.clicked.connect(self.export_data_to_excel)

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1186, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        # Connect button clicks to respective functions
        self.faceRecognitionBtn.clicked.connect(self.open_facial_recognition)
        self.registerStudentBtn.clicked.connect(self.open_register_student)
        self.studentMgmtBtn.clicked.connect(self.open_student_management)
        self.exitBtn.clicked.connect(QtWidgets.qApp.quit)
        self.exitBtn_2.clicked.connect(self.open_login_page)
        self.pushButton.clicked.connect(self.open_dashboard)
        self.pushButton2.clicked.connect(self.open_logs)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Personalized Authentication and Student Screening"))
        self.label_9.setText(_translate("MainWindow", "      PASS"))
        self.pushButton.setText(_translate("MainWindow", "Dashboard"))
        self.pushButton2.setText(_translate("MainWindow", "Logs"))
        self.faceRecognitionBtn.setText(_translate("MainWindow", "Facial Recognition"))
        self.registerStudentBtn.setText(
            _translate("MainWindow", "Student Registration")
        )
        self.studentMgmtBtn.setText(_translate("MainWindow", "Student Management"))
        self.exitBtn.setText(_translate("MainWindow", "Exit"))
        self.exitBtn_2.setText(_translate("MainWindow", "Logout"))
        # self.label.setText(
        #     _translate(
        #         "MainWindow",
        #         "PASS: Personalized Authentication and Student Surveillance",
        #     )
        # )
        self.label_10.setText(_translate("MainWindow", "Student Management"))

    def load_students(self):
        if self.search_has_input:

            return 
        
        connection = pymysql.connect(
            host="localhost", user="root", password="", db="pass_db"
        )
        self.connection = (
            connection  # Store the connection reference to access it later
        )
        cursor = connection.cursor()

        # Fetch student data from the database
        query = "SELECT id, first_name, middle_name, last_name, course, sr_code, gender FROM tbl_student where status = 'Active'"
        cursor.execute(query)
        students = cursor.fetchall()

        # Display students in the table
        row_count = len(students)
        column_count = 9  # Increase column count for edit and delete buttons

        self.tableWidget.setRowCount(row_count)
        self.tableWidget.setColumnCount(column_count)

        header_labels = [
            "First Name",
            "Middle Name",
            "Last Name",
            "Course",
            "SR Code",
            "Sex",
            "View Logs",
            "Edit",
            "Archive",
        ]
        self.tableWidget.setHorizontalHeaderLabels(header_labels)

        for i, student in enumerate(students):
            student_id = student[0]  # Get the student ID
            for j in range(column_count - 3):  # Adjust the range
                item = QTableWidgetItem(
                    str(student[j + 1])
                )  # Update the index of the student details
                self.tableWidget.setItem(i, j, item)  # Update the column index

            # Create and set the edit button
            edit_button = QPushButton("Edit")
            edit_button.clicked.connect(
                lambda checked, student_id=student_id: self.edit_student(student_id)
            )
            edit_button.setStyleSheet(
                '''
                QPushButton {
                    background-color: #007bff;  /* Bootstrap primary color */
                    border: none;
                    color: white;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-family: Arial;
                    font-size: 8pt;
                }

                QPushButton:hover {
                    background-color: #0069d9;  /* Darker shade on hover */
                }
                '''
            )


            self.tableWidget.setCellWidget(i, column_count - 2, edit_button)

            # Create and set the delete button
            delete_button = QPushButton("Archive")
            delete_button.clicked.connect(
                lambda checked, student_id=student_id: self.archive_student(student_id)
            )
            delete_button.setStyleSheet(
                '''
                QPushButton {
                    background-color: #dc3545;  
                    border: none;
                    color: white;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-family: Arial;
                    font-size: 8pt;
                }

                QPushButton:hover {
                    background-color: #c82333;  /* Darker shade on hover */
                }
                '''
            )
            self.tableWidget.setCellWidget(i, column_count - 1, delete_button)

            # Create and set the delete button
            view_logs_button = QPushButton("View Logs")
            view_logs_button.clicked.connect(
                lambda checked, student_id=student_id: self.view_logs(student_id)
            )
            view_logs_button.setStyleSheet(
                '''
                QPushButton {
                    background-color: #17a2b8;  /* Bootstrap info color */
                    border: none;
                    color: white;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-family: Arial;
                    font-size: 8pt;
                }

                QPushButton:hover {
                    background-color: #138496;  /* Darker shade on hover */
                }
                '''
            )
            self.tableWidget.setCellWidget(i, column_count - 3, view_logs_button)

        cursor.close()

        # # Schedule the next update after 1 second
        # QTimer.singleShot(1000, self.load_students)


    def search_logs(self, search_text, load_students=True):
        connection = pymysql.connect(
            host="localhost", user="root", password="", db="pass_db"
        )
        cursor = connection.cursor()

        if search_text:
            # Disable loading students when the search textbox has input
            self.search_has_input = True
        else:
            self.search_has_input = False
            QTimer.singleShot(1000, self.load_students)
            
        query = """
        SELECT id, first_name, middle_name, last_name, course, sr_code, gender
        FROM tbl_student
        WHERE status = 'Active' and first_name LIKE %s
            OR last_name LIKE %s
            OR sr_code LIKE %s
            OR course LIKE %s
        """
        search_pattern = f"%{search_text}%"  # Add wildcards for partial matching
        cursor.execute(query, (search_pattern, search_pattern, search_pattern, search_pattern))
        students = cursor.fetchall()

        row_count = len(students)
        column_count = 9  # Increase column count for edit and delete buttons

        self.tableWidget.setRowCount(row_count)
        self.tableWidget.setColumnCount(column_count)

        header_labels = [
            "First Name",
            "Middle Name",
            "Last Name",
            "Course",
            "SR Code",
            "Sex",
            "View Logs",
            "Edit",
            "Archive",
        ]
        self.tableWidget.setHorizontalHeaderLabels(header_labels)

        for i, student in enumerate(students):
            student_id = student[0]  # Get the student ID
            for j in range(column_count - 3):  # Adjust the range
                item = QTableWidgetItem(
                    str(student[j + 1])
                )  # Update the index of the student details
                self.tableWidget.setItem(i, j, item)  # Update the column index

            # Create and set the edit button
            edit_button = QPushButton("Edit")
            edit_button.clicked.connect(
                lambda checked, student_id=student_id: self.edit_student(student_id)
            )
            edit_button.setStyleSheet(
                '''
                QPushButton {
                    background-color: #007bff;  /* Bootstrap primary color */
                    border: none;
                    color: white;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-family: Arial;
                    font-size: 8pt;
                }

                QPushButton:hover {
                    background-color: #0069d9;  /* Darker shade on hover */
                }
                '''
            )
            self.tableWidget.setCellWidget(i, column_count - 2, edit_button)

            # Create and set the delete button
            delete_button = QPushButton("Archive")
            delete_button.clicked.connect(
                lambda checked, student_id=student_id: self.archive_student(student_id)
            )
            delete_button.setStyleSheet(
                '''
                QPushButton {
                    background-color: #dc3545;  
                    border: none;
                    color: white;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-family: Arial;
                    font-size: 8pt;
                }

                QPushButton:hover {
                    background-color: #c82333;  /* Darker shade on hover */
                }
                '''
            )
            self.tableWidget.setCellWidget(i, column_count - 1, delete_button)

            # Create and set the delete button
            view_logs_button = QPushButton("View Logs")
            view_logs_button.clicked.connect(
                lambda checked, student_id=student_id: self.view_logs(student_id)
            )
            view_logs_button.setStyleSheet(
                '''
                QPushButton {
                    background-color: #17a2b8;  /* Bootstrap info color */
                    border: none;
                    color: white;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-family: Arial;
                    font-size: 8pt;
                }

                QPushButton:hover {
                    background-color: #138496;  /* Darker shade on hover */
                }
                '''
            )
            self.tableWidget.setCellWidget(i, column_count - 3, view_logs_button)

        cursor.close()
        connection.close()

        if load_students:
            QTimer.singleShot(1000, self.load_students)

    def start_loading_students(self):
        # Start loading the students initially
        self.load_students()

    def view_logs(self, student_id):
        print("Edit Student ID:", student_id)
        dialog = ViewLogsDialog()
        dialog.setupUi(student_id)
        dialog.exec_()

    def edit_student(self, student_id):
        print("Edit Student ID:", student_id)
        dialog = EditStudentDialog()
        dialog.setupUi(student_id)
        dialog.exec_()

    def delete_student(self, student_id):
        print("Delete Student ID:", student_id)
        dialog = DeleteStudentDialog()
        dialog.setupUi(student_id)
        dialog.exec_()

    def archive_student(self, student_id):
        print("Archive Student ID:", student_id)
        dialog = DeleteStudentDialog()
        dialog.setupUi(student_id)
        dialog.exec_()


    def open_dashboard(self):
        print("Opening Dashboard...")
        # self.MainWindow.hide()
        # self.dashboard_window = QMainWindow()
        # self.ui = dashboard.Ui_Dashboard()
        # self.ui.setupUi(self.dashboard_window)
        # self.dashboard_window.show()

        self.MainWindow.hide()
        self.dashboard_window = QtWidgets.QMainWindow()
        self.ui = dashboard.Ui_Dashboard()
        self.ui.setupUi(self.dashboard_window)
        # self.ui.tableWidget.setParent(self.ui.centralwidget)
        # self.ui.load_logs()
        self.dashboard_window.show()



    def open_logs(self):
        print("Opening Logs...")
        self.MainWindow.hide()
        self.logs_window = QtWidgets.QMainWindow()
        self.ui = logs.Logs()
        self.ui.setupUi(self.logs_window)
        self.ui.tableWidget.setParent(self.ui.centralwidget)
        self.ui.load_logs()
        self.logs_window.show()

    def open_archives(self):
        print("Opening Archives...")
        self.MainWindow.hide()
        self.archives_window = QtWidgets.QMainWindow()
        self.ui = archive_management.ArchiveManagementWindow()
        self.ui.setupUi(self.archives_window)
        self.ui.tableWidget.setParent(self.ui.centralwidget)
        self.ui.load_students()
        self.archives_window.show()

    def open_facial_recognition(self):
        from facial_recognition import FacialRecognitionWindow

        face_recognition = FacialRecognitionWindow()
        # Add your code to open facial recognition here
        print("Opening Facial Recognition...")
        face_recognition.face_recognition_func()

    def open_register_student(self):
        # Add your code to open register student here
        print("Opening Register Student...")
        self.MainWindow.hide()
        self.register_student_window = QtWidgets.QMainWindow()
        self.ui = register_student.RegisterStudentWindow()
        self.ui.setupUi(self.register_student_window)
        self.register_student_window.show()

    def open_student_management(self):
        # Add your code to open student management here
        print("Opening Student Management Page...")
        self.MainWindow.hide()  # Hide the main window instead of closing it
        # Create the student management window
        self.student_management_window = QtWidgets.QMainWindow()
        self.ui = student_management.StudentManagementWindow()
        self.ui.setupUi(self.student_management_window)
        # Set the parent widget of the table widget
        self.ui.tableWidget.setParent(self.ui.centralwidget)
        # Load the students in the table
        self.ui.load_students()
        self.student_management_window.show()

    def open_login_page(self):
        # Add your code to open the login page here
        print("Opening Login Page...")
        self.MainWindow.hide()
        self.admin_login_window = QtWidgets.QMainWindow()
        self.ui = admin_login.Ui_MainWindow()
        self.ui.setupUi(self.admin_login_window)
        self.admin_login_window.show()

    def export_data_to_excel(self):
        # Connect to the MySQL database
        connection = pymysql.connect(
            host="localhost", user="root", password="", database="pass_db"
        )

        try:
            # Create a cursor object to execute SQL queries
            cursor = connection.cursor()

            # Retrieve data from the tbl_student table
            select_query = "SELECT first_name, middle_name, last_name, course, sr_code FROM tbl_student"
            cursor.execute(select_query)
            student_data = cursor.fetchall()

            # Create a new Excel workbook and select the active sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Write the column headers
            sheet["A1"] = "First Name"
            sheet["B1"] = "Middle Name"
            sheet["C1"] = "Last Name"
            sheet["D1"] = "Course"
            sheet["E1"] = "SR Code"

            for row_index, student in enumerate(student_data, start=2):
                sheet.cell(row=row_index, column=1).value = student[0]  # First Name
                sheet.cell(row=row_index, column=2).value = student[1]  # Middle Name
                sheet.cell(row=row_index, column=3).value = student[2]  # Last Name
                sheet.cell(row=row_index, column=4).value = student[3]  # Course
                sheet.cell(row=row_index, column=5).value = student[4]  # Course

            # Save the Excel file
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y-%m-%d_%H-%M-%S")
            self.file_name = f"student_data_{formatted_datetime}.xlsx"

            # Save the Excel file inside the "folder_path" folder
            self.file_path = rf"{self.folder_path}\{self.file_name}"

            # Create the "logs" folder if it doesn't exist
            if not os.path.exists(self.folder_path):
                os.makedirs(self.folder_path)
            # Save the Excel file
            workbook.save(self.file_path)
            print("Student Data exported to Excel successfully!")

        except Exception as e:
            print("Error exporting data to Excel:", str(e))

        finally:
            # Close the cursor and connection
            cursor.close()
            connection.close()

    # ...


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = StudentManagementWindow()
    ui.setupUi(MainWindow)
    ui.load_students()
    MainWindow.show()
    sys.exit(app.exec_())
