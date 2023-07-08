import sys
import os
import register_student
import student_management
import dashboard
import admin_login
import numpy as np
import pymysql
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (
    QTableWidget,
    QTableWidgetItem,
    QLabel,
)
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QPixmap
from about_dialog import AboutDialog
from PyQt5.QtCore import QTimer
from datetime import datetime, timedelta
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QTableWidget,
    QScrollArea,
)
import openpyxl
from datetime import datetime
import matplotlib.pyplot as plt
import auto_export_reset as auto


class Logs(object):
    def __init__(self) -> None:
        self.logs_sent = False
        self.existing = False

        # For Excel File
        self.file_name = ""
        self.file_path = ""
        self.folder_name = "Logs"
        self.folder_path = rf"C:\Users\SampleUser\Desktop\{self.folder_name}"  # Change this to your own file path

        # Export and Delete Analytics Every Monday (Check Every Second)
        self.day_of_export = auto.AutomaticExportResetLogs().day_of_export
        self.timer_second = 2
        self.timer = QTimer()
        self.timer.timeout.connect(self.analytics_reset_export_check)
        self.timer.start(self.timer_second * 1000)  # Execute every set self.time_second

    def analytics_reset_export_check(self):
        print(self.day_of_export)
        auto_export = auto.AutomaticExportResetLogs()

        current_date = datetime.now().date()
        prev_sunday = current_date - timedelta(days=1)
        prev_monday = current_date - timedelta(days=7)
        current_weekday = current_date.weekday()

        if auto.DAYS_IN_WEEK[current_weekday] == self.day_of_export:
            if auto_export.check_if_logs_found(
                prev_sunday=prev_sunday, prev_monday=prev_monday
            ):
                print("Exporting...")
                auto_export.export_logs(
                    prev_sunday=prev_sunday, prev_monday=prev_monday
                )
                auto_export.export_delete_analytics(
                    prev_sunday=prev_sunday, prev_monday=prev_monday
                )
                print("Exporting Done!")
            else:
                self.timer.stop()
                return
        else:
            print(f"Today is not {self.day_of_export}")
            self.timer.stop()
            return

    def setupUi(self, MainWindow):
        self.MainWindow = MainWindow
        MainWindow.setWindowFlags(
            MainWindow.windowFlags()
            & ~QtCore.Qt.WindowMinimizeButtonHint
            & ~QtCore.Qt.WindowMaximizeButtonHint
        )

        MainWindow.setObjectName("MainWindow")
        self.MainWindow.showMaximized()
        MainWindow.setWindowFlags(
            MainWindow.windowFlags()
            & ~QtCore.Qt.WindowCloseButtonHint  # Remove the close button
        )
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        # Add table widget
        self.tableWidget = QTableWidget(self.centralwidget)
        self.tableWidget.setGeometry(QtCore.QRect(310, 150, 1240, 450))
        self.tableWidget.setObjectName("tableWidget")

        # search bar
        self.searchLineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.searchLineEdit.setGeometry(QtCore.QRect(1180, 95, 200, 30))
        self.searchLineEdit.setObjectName("searchLineEdit")
        self.searchLineEdit.textChanged.connect(self.search_logs)

        self.search_has_input = False

        self.searchLabel = QtWidgets.QLabel(self.centralwidget)
        self.searchLabel.setGeometry(QtCore.QRect(1130, 95, 70, 30))
        self.searchLabel.setObjectName("searchLabel")
        self.searchLabel.setText("Search:")

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        self.frame = QtWidgets.QFrame(self.centralwidget)
        self.frame.setGeometry(QtCore.QRect(-1, 1, 261, 2000))
        self.frame.setStyleSheet("background-color: rgb(227, 30, 36);")
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.label_4 = QtWidgets.QLabel(self.frame)
        self.label_4.setGeometry(QtCore.QRect(80, 20, 101, 91))
        self.label_4.setText("")
        self.label_4.setPixmap(QtGui.QPixmap("img/logo2.png"))
        self.label_4.setScaledContents(True)
        self.label_4.setObjectName("label_4")
        self.label_9 = QtWidgets.QLabel(self.frame)
        self.label_9.setGeometry(QtCore.QRect(44, 110, 201, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label_9.setFont(font)
        self.label_9.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_9.setObjectName("label_9")

        self.dashboardBtn = QtWidgets.QPushButton(self.frame)
        self.dashboardBtn.setGeometry(QtCore.QRect(40, 200, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.dashboardBtn.setFont(font)
        self.dashboardBtn.setStyleSheet(
            " background-color: transparent;\n" "color: white;\n" ""
        )
        self.dashboardBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.dashboardBtn.setObjectName("dashboardBtn")

        self.logsBtn = QtWidgets.QPushButton(self.frame)
        self.logsBtn.setGeometry(QtCore.QRect(40, 250, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.logsBtn.setFont(font)
        self.logsBtn.setStyleSheet(
            " background-color: transparent;\n" "color: white;\n" ""
        )
        self.logsBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.logsBtn.setObjectName("logsBtn")

        self.faceRecognitionBtn = QtWidgets.QPushButton(self.frame)
        self.faceRecognitionBtn.setGeometry(QtCore.QRect(40, 300, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.faceRecognitionBtn.setFont(font)
        self.faceRecognitionBtn.setStyleSheet(
            " background-color: transparent;\n" "color: white;\n" ""
        )
        self.faceRecognitionBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.faceRecognitionBtn.setObjectName("faceRecognitionBtn")
        self.registerStudentBtn = QtWidgets.QPushButton(self.frame)
        self.registerStudentBtn.setGeometry(QtCore.QRect(40, 350, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.registerStudentBtn.setFont(font)
        self.registerStudentBtn.setStyleSheet(
            " background-color: transparent;\n" "color: white;\n" ""
        )
        self.registerStudentBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.registerStudentBtn.setObjectName("registerStudentBtn")
        self.studentMgmtBtn = QtWidgets.QPushButton(self.frame)
        self.studentMgmtBtn.setGeometry(QtCore.QRect(40, 400, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.studentMgmtBtn.setFont(font)
        self.studentMgmtBtn.setStyleSheet(
            " background-color: transparent;\n" "color: white;\n" ""
        )
        self.studentMgmtBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.studentMgmtBtn.setObjectName("studentMgmtBtn")
        self.exitBtn = QtWidgets.QPushButton(self.frame)
        self.exitBtn.setGeometry(QtCore.QRect(40, 450, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.exitBtn.setFont(font)
        self.exitBtn.setStyleSheet(
            " background-color: transparent;\n" "color: white;\n" ""
        )
        self.exitBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.exitBtn.setObjectName("exitBtn")
        self.exitBtn_2 = QtWidgets.QPushButton(self.frame)
        self.exitBtn_2.setGeometry(QtCore.QRect(40, 500, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.exitBtn_2.setFont(font)
        self.exitBtn_2.setStyleSheet(
            " background-color: transparent;\n" "color: white;\n" ""
        )
        self.exitBtn_2.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.exitBtn_2.setObjectName("exitBtn_2")

        self.aboutBtn = QtWidgets.QPushButton(self.frame)
        self.aboutBtn.setGeometry(QtCore.QRect(40, 550, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.aboutBtn.setFont(font)
        self.aboutBtn.setStyleSheet(
            " background-color: transparent;\n" "color: white;\n" ""
        )
        self.aboutBtn.setStyleSheet(
            " background-color: transparent;\n"
            "color: white;\n"
            "text-align: left;\n"
            ""
        )
        self.aboutBtn.setObjectName("aboutBtn")

        # self.frame_3 = QtWidgets.QFrame(self.centralwidget)
        # self.frame_3.setGeometry(QtCore.QRect(261, -1, 2000, 61))
        # self.frame_3.setStyleSheet("background-color: rgb(255, 255, 255);")
        # self.frame_3.setFrameShape(QtWidgets.QFrame.StyledPanel)
        # self.frame_3.setFrameShadow(QtWidgets.QFrame.Raised)
        # self.frame_3.setObjectName("frame_3")
        # self.label = QtWidgets.QLabel(self.frame_3)
        # self.label.setGeometry(QtCore.QRect(20, 10, 671, 41))
        # font = QtGui.QFont()
        # font.setFamily("Arial")
        # font.setPointSize(12)
        # self.label.setFont(font)
        # self.label.setObjectName("label")
        self.label_10 = QtWidgets.QLabel(self.centralwidget)
        self.label_10.setGeometry(QtCore.QRect(310, 90, 201, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(15)
        font.setBold(True)
        font.setWeight(75)
        self.label_10.setFont(font)
        self.label_10.setStyleSheet("color:black;")
        self.label_10.setObjectName("label_10")

        self.exportDataBtn = QtWidgets.QPushButton(self.centralwidget)
        self.exportDataBtn.setGeometry(QtCore.QRect(1420, 90, 121, 40))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(8)
        self.exportDataBtn.setFont(font)
        self.exportDataBtn.setObjectName("exportDataBtn")
        self.exportDataBtn.setText("Export Data")
        self.exportDataBtn.setStyleSheet(
            """
            QPushButton {
                background-color: #dc3545;  
                border: none;
                color: white;
                padding: 8px 16px;
                border-radius: 4px;
                font-family: Arial;
                font-size: 8pt;
            }

            QPushButton:hover {
                background-color: #c82333;  
            }
            """
        )
        self.exportDataBtn.clicked.connect(self.export_data_to_excel)

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1186, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        # Connect button clicks to respective functions
        self.faceRecognitionBtn.clicked.connect(self.open_facial_recognition)
        self.registerStudentBtn.clicked.connect(self.open_register_student)
        self.studentMgmtBtn.clicked.connect(self.open_student_management)
        self.exitBtn.clicked.connect(QtWidgets.qApp.quit)
        self.exitBtn_2.clicked.connect(self.open_login_page)
        self.aboutBtn.clicked.connect(self.open_about)
        self.dashboardBtn.clicked.connect(self.open_dashboard)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(
            _translate(
                "MainWindow",
                "PASS: Personalized Authentication and Student Surveillance",
            )
        )
        self.label_9.setText(_translate("MainWindow", "      PASS"))
        self.dashboardBtn.setText(_translate("MainWindow", "Dashboard"))
        self.logsBtn.setText(_translate("MainWindow", "Logs"))
        self.faceRecognitionBtn.setText(_translate("MainWindow", "Facial Recognition"))
        self.registerStudentBtn.setText(
            _translate("MainWindow", "Student Registration")
        )
        self.studentMgmtBtn.setText(_translate("MainWindow", "Student Management"))
        self.aboutBtn.setText(_translate("MainWindow", "About System"))
        self.exitBtn.setText(_translate("MainWindow", "Exit"))
        self.exitBtn_2.setText(_translate("MainWindow", "Logout"))
        # self.label.setText(
        #     _translate(
        #         "MainWindow",
        #         "PASS: Personalized Authentication and Student Surveillance",
        #     )
        # )
        self.label_10.setText(_translate("MainWindow", "Logs"))

    def load_logs(self):
        if self.search_has_input:
            return

        connection = pymysql.connect(
            host="localhost", user="root", password="", db="pass_db"
        )
        cursor = connection.cursor()

        delete_query = "DELETE FROM tbl_logs WHERE date_log < %s"
        week_ago = datetime.now() - timedelta(days=7)
        cursor.execute(delete_query, (week_ago.date(),))
        connection.commit()

        query = "SELECT  tbl_student.image, tbl_student.first_name, tbl_student.last_name, tbl_student.course, tbl_student.sr_code, tbl_student.gender, tbl_logs.date_log, tbl_logs.time_log, tbl_logs.log_type FROM tbl_logs LEFT JOIN tbl_student ON tbl_logs.student_id = tbl_student.id"
        cursor.execute(query)
        logs = cursor.fetchall()

        row_count = len(logs)
        column_count = 9  # Increase the column count for the image column

        self.tableWidget.setRowCount(row_count)
        self.tableWidget.setColumnCount(column_count)

        header_labels = [
            "Image",
            "First Name",
            "Last Name",
            "Course",
            "SR Code",
            "Gender",
            "Date Log",
            "Time Log",
            "Log Type",
        ]
        self.tableWidget.setHorizontalHeaderLabels(header_labels)

        for row, log in enumerate(logs):
            (
                image_filename,
                first_name,
                last_name,
                course,
                sr_code,
                gender,
                date_log,
                time_log,
                log_type,
            ) = log

            # Create a QLabel and set the image pixmap
            image_label = QLabel()
            image_path = os.path.join("images", str(image_filename))
            pixmap = QPixmap(image_path)
            if not pixmap.isNull():
                # Calculate the size of the cell
                cell_width = self.tableWidget.columnWidth(0)
                cell_height = self.tableWidget.rowHeight(row)

                # Resize the pixmap to fit the cell dimensions
                scaled_pixmap = pixmap.scaled(
                    cell_width, cell_height, Qt.AspectRatioMode.KeepAspectRatio
                )

                # Set the scaled pixmap on the image label
                image_label.setPixmap(scaled_pixmap)
                image_label.setAlignment(
                    Qt.AlignCenter
                )  # Center the image in the label

            self.tableWidget.setCellWidget(row, 0, image_label)

            self.tableWidget.setItem(row, 1, QTableWidgetItem(first_name))
            self.tableWidget.setItem(row, 2, QTableWidgetItem(last_name))
            self.tableWidget.setItem(row, 3, QTableWidgetItem(course))
            self.tableWidget.setItem(row, 4, QTableWidgetItem(sr_code))
            self.tableWidget.setItem(row, 5, QTableWidgetItem(gender))
            self.tableWidget.setItem(row, 6, QTableWidgetItem(str(date_log)))
            self.tableWidget.setItem(row, 7, QTableWidgetItem(str(time_log)))
            self.tableWidget.setItem(row, 8, QTableWidgetItem(log_type))

        cursor.close()

        # Schedule the next update after 1 second
        # QTimer.singleShot(1000, self.load_logs)

    def search_logs(self, search_text):
        connection = pymysql.connect(
            host="localhost", user="root", password="", db="pass_db"
        )
        cursor = connection.cursor()

        if search_text:
            self.search_has_input = True
        else:
            self.search_has_input = False
            QTimer.singleShot(1000, self.load_logs)

        query = """
        SELECT tbl_student.image, tbl_student.first_name, tbl_student.last_name,
               tbl_student.course, tbl_student.sr_code, tbl_student.gender, tbl_logs.date_log, tbl_logs.time_log, tbl_logs.log_type
        FROM tbl_logs
        LEFT JOIN tbl_student ON tbl_logs.student_id = tbl_student.id
        WHERE tbl_student.first_name LIKE %s
            OR tbl_student.last_name LIKE %s
            OR tbl_student.sr_code LIKE %s
            OR tbl_student.course LIKE %s
        """
        search_pattern = f"%{search_text}%"  # Add wildcards for partial matching
        cursor.execute(
            query, (search_pattern, search_pattern, search_pattern, search_pattern)
        )
        logs = cursor.fetchall()

        row_count = len(logs)
        column_count = 9  # Increase the column count for the image column

        self.tableWidget.setRowCount(row_count)
        self.tableWidget.setColumnCount(column_count)

        header_labels = [
            "Image",
            "First Name",
            "Last Name",
            "Course",
            "SR Code",
            "Gender",
            "Date Log",
            "Time Log",
            "Log Type",
        ]
        self.tableWidget.setHorizontalHeaderLabels(header_labels)

        for row, log in enumerate(logs):
            (
                image_filename,
                first_name,
                last_name,
                course,
                sr_code,
                gender,
                date_log,
                time_log,
                log_type,
            ) = log

            # Create a QLabel and set the image pixmap
            image_label = QLabel()
            image_path = os.path.join("images", str(image_filename))
            pixmap = QPixmap(image_path)
            if not pixmap.isNull():
                # Calculate the size of the cell
                cell_width = self.tableWidget.columnWidth(0)
                cell_height = self.tableWidget.rowHeight(row)

                # Resize the pixmap to fit the cell dimensions
                scaled_pixmap = pixmap.scaled(
                    cell_width, cell_height, Qt.AspectRatioMode.KeepAspectRatio
                )

                # Set the scaled pixmap on the image label
                image_label.setPixmap(scaled_pixmap)
                image_label.setAlignment(
                    Qt.AlignCenter
                )  # Center the image in the label

            self.tableWidget.setCellWidget(row, 0, image_label)

            self.tableWidget.setItem(row, 1, QTableWidgetItem(first_name))
            self.tableWidget.setItem(row, 2, QTableWidgetItem(last_name))
            self.tableWidget.setItem(row, 3, QTableWidgetItem(course))
            self.tableWidget.setItem(row, 4, QTableWidgetItem(sr_code))
            self.tableWidget.setItem(row, 5, QTableWidgetItem(gender))
            self.tableWidget.setItem(row, 6, QTableWidgetItem(str(date_log)))
            self.tableWidget.setItem(row, 7, QTableWidgetItem(str(time_log)))
            self.tableWidget.setItem(row, 8, QTableWidgetItem(log_type))

        cursor.close()
        connection.close()

    def start_loading_students(self):
        # Start loading the students initially
        self.load_logs()

    def open_facial_recognition(self):
        from facial_recognition import FacialRecognitionWindow

        face_recognition = FacialRecognitionWindow()
        # Add your code to open facial recognition here
        print("Opening Facial Recognition...")
        face_recognition.face_recognition_func()

    def open_about(self):
        print("Opening About System Dialog")
        dialog = AboutDialog()
        dialog.setupUi()
        dialog.show()

    def open_register_student(self):
        print("Opening Register Student...")
        self.MainWindow.hide()
        self.register_student_window = QtWidgets.QMainWindow()
        self.ui = register_student.RegisterStudentWindow()
        self.ui.setupUi(self.register_student_window)
        self.register_student_window.show()

    def open_dashboard(self):
        print("Opening Dashboard...")
        self.MainWindow.hide()
        self.dashboard_window = QtWidgets.QMainWindow()
        self.ui = dashboard.Ui_Dashboard()
        self.ui.setupUi(self.dashboard_window)
        self.dashboard_window.show()

    def open_student_management(self):
        print("Opening Student Management Page...")
        self.MainWindow.hide()
        self.student_management_window = QtWidgets.QMainWindow()
        self.ui = student_management.StudentManagementWindow()
        self.ui.setupUi(self.student_management_window)
        # Set the parent widget of the table widget
        self.ui.tableWidget.setParent(self.ui.centralwidget)
        # Load the students in the table
        self.ui.load_students()
        self.student_management_window.show()

    def open_login_page(self):
        print("Opening Login Page...")
        self.MainWindow.hide()
        self.admin_login_window = QtWidgets.QMainWindow()
        self.ui = admin_login.Ui_MainWindow()
        self.ui.setupUi(self.admin_login_window)
        self.admin_login_window.show()

    def export_data_to_excel(self):
        # Connect to the MySQL database
        connection = pymysql.connect(
            host="localhost", user="root", password="", database="pass_db"
        )

        try:
            # Create a cursor object to execute SQL queries
            cursor = connection.cursor()

            # Retrieve data from the tbl_student table
            select_query = "SELECT CONCAT(tbl_student.first_name, ' ', tbl_student.last_name) AS student_name, tbl_student.course, tbl_student.sr_code, tbl_logs.date_log, tbl_logs.time_log FROM tbl_logs LEFT JOIN tbl_student ON tbl_logs.student_id = tbl_student.id"
            cursor.execute(select_query)
            student_data = cursor.fetchall()

            # Create a new Excel workbook and select the active sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Write the column headers
            sheet["A1"] = "Student Name"
            sheet["B1"] = "Course"
            sheet["C1"] = "SR Code"
            sheet["D1"] = "Date Log"
            sheet["E1"] = "Time Log"

            # Set column width for date columns
            date_columns = ["D", "E"]  # Columns D and E represent the date columns
            for column in date_columns:
                sheet.column_dimensions[
                    column
                ].width = 15  # Adjust the width as per your preference

            for row_index, student in enumerate(student_data, start=2):
                sheet.cell(row=row_index, column=1).value = student[0]
                sheet.cell(row=row_index, column=2).value = student[1]
                sheet.cell(row=row_index, column=3).value = student[2]
                sheet.cell(row=row_index, column=4).value = student[3]
                sheet.cell(row=row_index, column=5).value = student[4]

            # Save the Excel file
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y-%m-%d_%H-%M-%S")
            self.file_name = f"logs_data_{formatted_datetime}.xlsx"

            # Save the Excel file inside the "folder_path" folder
            self.file_path = rf"{self.folder_path}\{self.file_name}"

            # Create the "logs" folder if it doesn't exist
            if not os.path.exists(self.folder_path):
                os.makedirs(self.folder_path)
            # Save the Excel file
            workbook.save(self.file_path)
            print("Logs Data exported to Excel successfully!")

        except Exception as e:
            print("Error exporting data to Excel:", str(e))

        finally:
            # Close the cursor and connection
            cursor.close()
            connection.close()


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Logs()
    ui.setupUi(MainWindow)
    ui.load_logs()
    MainWindow.show()
    sys.exit(app.exec_())
